"""
主窗口 - 剑网三副本工资统计
手动保存 + 5分钟自动保存 + 保存后可撤销/重做（30步） + 窗口高度自适应 + 运算式行变灰（右键开关）
修复：缓存新增行编辑后表达式未同步导致无法变灰
"""

import os
import datetime
import json

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QComboBox, QTableView, QHeaderView,
    QFileDialog, QLabel, QMenu
)
from PyQt6.QtCore import Qt, QDate, QTimer, QPoint
from PyQt6.QtGui import QFont, QShortcut, QKeySequence, QAction

from core.database import (
    init_database, get_all_records, get_date_list,
    add_record, update_record, delete_record, get_connection
)
from core.importer import import_excel
from ui.table_model import SalaryTableModel, NoBorderDelegate
from ui.dialogs import RecordDialog, show_message, confirm_action
from core.utils import number_to_brick
from openpyxl import Workbook

SETTINGS_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'settings.json')


def _calc_stats(records):
    if not records:
        return {
            'count': 0,
            'avg_normal_salary': 0, 'avg_normal_consume': 0,
            'avg_hero_salary': 0, 'avg_hero_consume': 0, 'avg_total_salary': 0,
            'sum_normal_salary': 0, 'sum_normal_consume': 0,
            'sum_hero_salary': 0, 'sum_hero_consume': 0, 'sum_total_salary': 0,
        }
    n = len(records)
    sums = [sum(r[i] for r in records) for i in [4, 5, 6, 7, 8]]
    return {
        'count': n,
        'avg_normal_salary': sums[0] // n, 'avg_normal_consume': sums[1] // n,
        'avg_hero_salary': sums[2] // n, 'avg_hero_consume': sums[3] // n,
        'avg_total_salary': sums[4] // n,
        'sum_normal_salary': sums[0], 'sum_normal_consume': sums[1],
        'sum_hero_salary': sums[2], 'sum_hero_consume': sums[3],
        'sum_total_salary': sums[4],
    }


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('剑网三 副本工资统计')
        self.setMinimumWidth(1200)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)

        self.pending_records = []
        self.pending_deletes = []
        self.pending_edits = {}

        self.undo_stack = []
        self.redo_stack = []
        self.max_history = 30

        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.setSingleShot(True)
        self.auto_save_timer.timeout.connect(self._auto_save)

        self.settings = self._load_settings()
        self.visible_columns = self.settings.get('visible_columns', list(SalaryTableModel.HEADERS))
        self.column_order = self.settings.get('column_order', list(SalaryTableModel.HEADERS))
        self.gray_expression_rows = self.settings.get('gray_expression_rows', True)

        init_database()
        self._init_ui()
        self._set_default_week()
        self._apply_column_settings()

        self.table_model.set_gray_expression_rows(self.gray_expression_rows)

        self._refresh_data()

        self.save_shortcut = QShortcut(QKeySequence('Ctrl+S'), self)
        self.save_shortcut.activated.connect(self._on_save)
        self.undo_shortcut = QShortcut(QKeySequence('Ctrl+Z'), self)
        self.undo_shortcut.activated.connect(self._on_undo)
        self.redo_shortcut = QShortcut(QKeySequence('Ctrl+Y'), self)
        self.redo_shortcut.activated.connect(self._on_redo)

    # ---------- UI 构建 ----------
    def _init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        title_bar = QWidget()
        title_bar.setObjectName('titleBar')
        title_bar.setFixedHeight(40)
        title_bar_layout = QHBoxLayout(title_bar)
        title_bar_layout.setContentsMargins(16, 0, 8, 0)
        title_bar_layout.setSpacing(0)
        bar_title = QLabel('JX3 SALARY STATS')
        bar_title.setObjectName('barTitle')
        title_bar_layout.addWidget(bar_title)
        title_bar_layout.addStretch()
        min_btn = QPushButton('_')
        min_btn.setObjectName('minBtn')
        min_btn.setFixedSize(36, 28)
        min_btn.clicked.connect(self.showMinimized)
        title_bar_layout.addWidget(min_btn)
        close_btn = QPushButton('X')
        close_btn.setObjectName('closeBtn')
        close_btn.setFixedSize(36, 28)
        close_btn.clicked.connect(self.close)
        title_bar_layout.addWidget(close_btn)
        main_layout.addWidget(title_bar)

        self.content_widget = QWidget()
        content_layout = QVBoxLayout(self.content_widget)
        content_layout.setSpacing(8)
        content_layout.setContentsMargins(10, 10, 10, 10)

        toolbar_layout = QHBoxLayout()
        title_label = QLabel('剑网三 副本工资统计')
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setObjectName('titleLabel')
        toolbar_layout.addWidget(title_label)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(QLabel('日期:'))
        self.date_filter_combo = QComboBox()
        self.date_filter_combo.setMinimumWidth(150)
        self.date_filter_combo.currentTextChanged.connect(self._on_date_filter_changed)
        toolbar_layout.addWidget(self.date_filter_combo)
        toolbar_layout.addWidget(QLabel('门派:'))
        self.faction_filter_combo = QComboBox()
        self.faction_filter_combo.setMinimumWidth(100)
        self.faction_filter_combo.addItem('全部')
        self.faction_filter_combo.currentTextChanged.connect(self._filter_data)
        toolbar_layout.addWidget(self.faction_filter_combo)
        toolbar_layout.addStretch()

        self.import_btn = QPushButton('导入Excel')
        self.import_btn.clicked.connect(self._on_import)
        toolbar_layout.addWidget(self.import_btn)
        self.export_btn = QPushButton('导出Excel')
        self.export_btn.clicked.connect(self._on_export)
        toolbar_layout.addWidget(self.export_btn)
        self.add_btn = QPushButton('新增记录')
        self.add_btn.clicked.connect(self._on_add)
        toolbar_layout.addWidget(self.add_btn)
        self.delete_btn = QPushButton('删除')
        self.delete_btn.clicked.connect(self._on_delete)
        toolbar_layout.addWidget(self.delete_btn)
        self.save_btn = QPushButton('保存')
        self.save_btn.setObjectName('saveBtn')
        self.save_btn.clicked.connect(self._on_save)
        toolbar_layout.addWidget(self.save_btn)

        self.column_menu_btn = QPushButton('列设置')
        self.column_menu_btn.clicked.connect(self._show_column_menu)
        toolbar_layout.addWidget(self.column_menu_btn)

        content_layout.addLayout(toolbar_layout)

        self.table_view = QTableView()
        self.table_view.setAlternatingRowColors(False)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setFocusPolicy(Qt.FocusPolicy.NoFocus)

        self.table_view.horizontalHeader().setSectionsMovable(True)
        self.table_view.horizontalHeader().sectionMoved.connect(self._on_section_moved)

        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self._on_table_context_menu)

        self.table_model = SalaryTableModel()
        self.table_model.dataChanged.connect(self._on_cell_changed)
        self.table_view.setModel(self.table_model)
        self.table_view.setItemDelegate(NoBorderDelegate(self.table_view))
        content_layout.addWidget(self.table_view)
        main_layout.addWidget(self.content_widget)

        self.status_label = QLabel('就绪')
        self.statusBar().addWidget(self.status_label)
        self._drag_pos = None

    # ---------- 列设置 ----------
    def _apply_column_settings(self):
        order = self.column_order if self.column_order else SalaryTableModel.HEADERS
        ordered_visible = [name for name in order if name in self.visible_columns and name in SalaryTableModel.HEADERS]
        for name in SalaryTableModel.HEADERS:
            if name not in ordered_visible and name in self.visible_columns:
                ordered_visible.append(name)
        self.visible_columns = ordered_visible
        self.table_model.set_visible_columns(self.visible_columns)

    def _show_column_menu(self):
        menu = QMenu(self)
        for name in SalaryTableModel.HEADERS:
            action = QAction(name, menu, checkable=True)
            action.setChecked(name in self.visible_columns)
            action.toggled.connect(lambda checked, n=name: self._toggle_column(n, checked))
            menu.addAction(action)
        menu.exec(self.column_menu_btn.mapToGlobal(self.column_menu_btn.rect().bottomLeft()))

    def _toggle_column(self, col_name, visible):
        if visible and col_name not in self.visible_columns:
            self.visible_columns.append(col_name)
            order = self.column_order if self.column_order else SalaryTableModel.HEADERS
            self.visible_columns.sort(key=lambda x: order.index(x) if x in order else len(order))
        elif not visible and col_name in self.visible_columns:
            self.visible_columns.remove(col_name)
        self.table_model.set_visible_columns(self.visible_columns)
        self._save_settings()
        self._refresh_data()

    def _on_section_moved(self, logicalIndex, oldVisualIndex, newVisualIndex):
        header = self.table_view.horizontalHeader()
        visual_order = []
        for visual in range(header.count()):
            logical = header.logicalIndex(visual)
            visual_order.append(self.table_model.visible_columns[logical])
        self.table_model.visible_columns = visual_order
        self.column_order = visual_order
        self._save_settings()

    def _save_settings(self):
        try:
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump({
                    'visible_columns': self.visible_columns,
                    'column_order': self.column_order,
                    'gray_expression_rows': self.gray_expression_rows
                }, f)
        except:
            pass

    def _load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {}

    # ---------- 窗口拖拽 ----------
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None

    def closeEvent(self, event):
        unsaved = len(self.pending_records) + len(self.pending_deletes) + len(self.pending_edits)
        if unsaved > 0:
            ok = confirm_action(self, '未保存的更改', f'有 {unsaved} 条未保存的更改\n\n确定不保存就关闭吗？')
            if not ok:
                event.ignore()
                return
        event.accept()

    def _set_default_week(self):
        utc_now = datetime.datetime.utcnow()
        beijing_now = utc_now + datetime.timedelta(hours=8)
        weekday = beijing_now.weekday()
        hour = beijing_now.hour
        if weekday == 0 and hour < 7:
            monday = beijing_now.date() - datetime.timedelta(days=7)
        else:
            monday = beijing_now.date() - datetime.timedelta(days=weekday)
        sunday = monday + datetime.timedelta(days=6)
        start_str = f'{monday.month:02d}.{monday.day:02d}'
        end_str = f'{sunday.month:02d}.{sunday.day:02d}'
        self._default_week_str = f'{start_str}-{end_str}'
        self._first_load = True

    def _copy_last_week_characters(self):
        date_list = get_date_list()
        if not date_list:
            return
        last_week = None
        for d in reversed(date_list):
            if d != self._default_week_str:
                last_week = d
                break
        if not last_week:
            return
        last_week_records = get_all_records(last_week)
        if not last_week_records:
            return
        existing_names = set()
        current_records = get_all_records(self._default_week_str)
        for r in current_records:
            existing_names.add(r[2])
        for pr in self.pending_records:
            if pr[0] == self._default_week_str:
                existing_names.add(pr[3])
        parts = self._default_week_str.split('-')
        start_date = parts[0]
        end_date = parts[1] if len(parts) > 1 else parts[0]
        for r in last_week_records:
            name = r[2]
            if name not in existing_names:
                new_record = (self._default_week_str, start_date, end_date, name, r[3], 0, 0, 0, 0, '', '')
                self.pending_records.append(new_record)
                existing_names.add(name)
        if self.pending_records:
            self._reset_auto_save_timer()

    # ---------- 数据刷新 ----------
    def _refresh_data(self):
        if hasattr(self, '_first_load') and self._first_load:
            self._first_load = False
            idx = self.date_filter_combo.findText(self._default_week_str)
            if idx >= 0:
                self.date_filter_combo.setCurrentIndex(idx)
            else:
                self.date_filter_combo.addItem(self._default_week_str)
                self.date_filter_combo.setCurrentText(self._default_week_str)
                self._copy_last_week_characters()

        current_date = self.date_filter_combo.currentText()
        date_filter = current_date if current_date and current_date != '全部' else None

        all_db = get_all_records(None)

        base_records = []
        for r in all_db:
            if r[0] not in self.pending_deletes:
                if r[0] in self.pending_edits:
                    base_records.append(self.pending_edits[r[0]])
                else:
                    base_records.append(r)

        # 合并缓存新增 - 优先使用模型记录（保留表达式）
        for i, pr in enumerate(self.pending_records):
            temp_id = -(i + 1)
            existing = None
            for rec in self.table_model.records:
                if rec[0] == temp_id:
                    existing = rec
                    break
            if existing and len(existing) >= 15:
                temp_record = existing
            else:
                total = pr[5] + pr[7] - pr[6] - pr[8]
                temp_record = (temp_id, pr[0], pr[3], pr[4], pr[5], pr[6], pr[7], pr[8], total, pr[9], pr[10])
            if not date_filter or pr[0] == date_filter:
                base_records.append(temp_record)

        if date_filter:
            base_records = [r for r in base_records if r[1] == date_filter]

        if date_filter is None:
            aggregated = {}
            for r in base_records:
                name = r[2]
                if name in ('平均', '合计'):
                    continue
                if name not in aggregated:
                    aggregated[name] = [r[3], 0, 0, 0, 0]
                agg = aggregated[name]
                agg[1] += r[4]
                agg[2] += r[5]
                agg[3] += r[6]
                agg[4] += r[7]
            base_records = []
            for i, (name, data) in enumerate(aggregated.items()):
                faction = data[0]
                total = data[1] + data[3] - data[2] - data[4]
                base_records.append((-i - 1, '', name, faction, data[1], data[2], data[3], data[4], total, '', ''))
            base_records.sort(key=lambda r: r[2])
            self.table_model.set_editable_columns([])
            self.add_btn.setEnabled(False)
            self.delete_btn.setEnabled(False)
        else:
            self.table_model.set_editable_columns(['角色名', '门派', '普通工资', '普通消费', '英雄工资', '英雄消费', '团牌', '掉落'])
            self.add_btn.setEnabled(True)
            self.delete_btn.setEnabled(True)

        current_faction = self.faction_filter_combo.currentText()
        self.faction_filter_combo.blockSignals(True)
        self.faction_filter_combo.clear()
        self.faction_filter_combo.addItem('全部')
        factions = sorted({r[3] for r in all_db})
        for f in factions:
            self.faction_filter_combo.addItem(f)
        idx = self.faction_filter_combo.findText(current_faction)
        if idx >= 0:
            self.faction_filter_combo.setCurrentIndex(idx)
        else:
            self.faction_filter_combo.setCurrentIndex(0)
        self.faction_filter_combo.blockSignals(False)

        self.date_filter_combo.blockSignals(True)
        self.date_filter_combo.clear()
        self.date_filter_combo.addItem('全部')
        for d in get_date_list():
            self.date_filter_combo.addItem(d)
        if current_date and current_date != '全部' and self.date_filter_combo.findText(current_date) == -1:
            self.date_filter_combo.addItem(current_date)
        self.date_filter_combo.setCurrentText(current_date)
        self.date_filter_combo.blockSignals(False)

        faction_filter = self.faction_filter_combo.currentText()
        if faction_filter != '全部':
            base_records = [r for r in base_records if r[3] == faction_filter]

        data_records = [r for r in base_records if r[2] not in ('平均', '合计')]
        stats = _calc_stats(data_records)
        self.table_model.load_data(base_records, stats)

        header = self.table_view.horizontalHeader()
        for i in range(header.count()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

        window_title = '剑网三 副本工资统计'
        if current_date and current_date != '全部':
            window_title += f' - {current_date}'
        else:
            window_title += ' - 全部日期'
        self.setWindowTitle(window_title)

        unsaved = len(self.pending_records) + len(self.pending_deletes) + len(self.pending_edits)
        if unsaved > 0:
            parts = []
            if self.pending_records:
                parts.append(f'{len(self.pending_records)} 新增')
            if self.pending_deletes:
                parts.append(f'{len(self.pending_deletes)} 删除')
            if self.pending_edits:
                parts.append(f'{len(self.pending_edits)} 编辑')
            self.status_label.setText(f'共 {len(data_records)} 条记录 | 未保存: {", ".join(parts)}')
            self.save_btn.setEnabled(True)
        else:
            self.status_label.setText(f'共 {len(data_records)} 条记录')
            self.save_btn.setEnabled(False)

        # 强制刷新灰度状态
        self.table_model.set_gray_expression_rows(self.gray_expression_rows)

        self._adjust_window_height(len(data_records))

    def _adjust_window_height(self, data_count):
        if self.isMaximized():
            return
        header_height = self.table_view.horizontalHeader().height()
        row_height = self.table_view.rowHeight(0) if data_count > 0 else 30
        stats_rows = 2 if self.table_model.show_stats and data_count > 0 else 0
        total_rows = data_count + stats_rows
        table_height = header_height + row_height * total_rows + 4
        fixed_height = 40 + 55 + 30 + 20
        ideal_height = table_height + fixed_height
        screen = self.screen().availableGeometry()
        max_height = int(screen.height() * 0.9)
        ideal_height = min(ideal_height, max_height)
        ideal_height = max(ideal_height, 400)
        self.resize(self.width(), ideal_height)

    def _filter_data(self):
        self._refresh_data()

    def _on_date_filter_changed(self):
        self._refresh_data()

    def _push_undo(self, item):
        self.undo_stack.append(item)
        if len(self.undo_stack) > self.max_history:
            self.undo_stack.pop(0)

    # ---------- 单元格编辑（修复缓存行表达式同步） ----------
    def _on_cell_changed(self, topLeft, bottomRight):
        row, col = topLeft.row(), topLeft.column()
        record_id = self.table_model.get_record_id(row)
        if record_id is None:
            return
        record = None
        for r in self.table_model.records:
            if r[0] == record_id:
                record = r
                break
        if record is None:
            return

        if record_id < 0:
            idx = -record_id - 1
            if idx < len(self.pending_records):
                pr_list = list(self.pending_records[idx])
                old_pr = tuple(pr_list)
                col_name = self.table_model.visible_columns[col]
                field_map = {'角色名': 3, '门派': 4, '普通工资': 5, '普通消费': 6,
                             '英雄工资': 7, '英雄消费': 8, '团牌': 9, '掉落': 10}
                if col_name in field_map:
                    new_val = record[SalaryTableModel.FIELD_MAP[col_name]]
                    pr_list[field_map[col_name]] = new_val
                    self.pending_records[idx] = tuple(pr_list)

                    # 同步表达式到模型记录
                    expr_map = {'普通工资': 11, '普通消费': 12, '英雄工资': 13, '英雄消费': 14}
                    if col_name in expr_map:
                        expr_idx = expr_map[col_name]
                        # 从模型记录中获取原始输入并写入 pending_records 的表达式占位？pending_records 没有表达式字段。
                        # 但是模型记录已经有了表达式（因为 setData 已更新），所以不用额外操作。
                        pass

                    self._push_undo({
                        'type': 'edit_pending',
                        'pending_idx': idx,
                        'old_data': old_pr,
                        'new_data': tuple(pr_list)
                    })
                    self.redo_stack.clear()
                self._update_stats()
                self._reset_auto_save_timer()
            return

        new_record = tuple(record)
        old_record = self.pending_edits.get(record_id, None)
        if old_record is None:
            all_db = get_all_records(None)
            for r in all_db:
                if r[0] == record_id:
                    old_record = r
                    break
        if old_record is None:
            return

        self.pending_edits[record_id] = new_record
        self._push_undo({
            'type': 'edit',
            'record_id': record_id,
            'old_record': old_record,
            'new_record': new_record
        })
        self.redo_stack.clear()
        self._update_stats()
        self._reset_auto_save_timer()

    def _update_stats(self):
        data_records = [r for r in self.table_model.records if r[2] not in ('平均', '合计')]
        stats = _calc_stats(data_records)
        self.table_model.statistics = stats
        if self.table_model.show_stats and stats['count'] > 0:
            last_row = self.table_model.rowCount() - 1
            self.table_model.dataChanged.emit(
                self.table_model.index(last_row - 1, 0),
                self.table_model.index(last_row, self.table_model.columnCount() - 1),
                [Qt.ItemDataRole.DisplayRole]
            )
        unsaved = len(self.pending_records) + len(self.pending_deletes) + len(self.pending_edits)
        count = len(data_records)
        if unsaved > 0:
            parts = []
            if self.pending_records:
                parts.append(f'{len(self.pending_records)} 新增')
            if self.pending_deletes:
                parts.append(f'{len(self.pending_deletes)} 删除')
            if self.pending_edits:
                parts.append(f'{len(self.pending_edits)} 编辑')
            self.status_label.setText(f'共 {count} 条记录 | 未保存: {", ".join(parts)}')
            self.save_btn.setEnabled(True)
        else:
            self.status_label.setText(f'共 {count} 条记录')
            self.save_btn.setEnabled(False)

    def _on_add(self):
        dialog = RecordDialog(self)
        if dialog.exec() == RecordDialog.DialogCode.Accepted:
            data = dialog.get_data()
            self.pending_records.append(data)
            idx = len(self.pending_records) - 1
            self._push_undo({
                'type': 'add',
                'index': idx,
                'data': data
            })
            self.redo_stack.clear()
            self._refresh_data()
            self._reset_auto_save_timer()

    def _on_delete(self):
        index = self.table_view.currentIndex()
        if not index.isValid():
            show_message(self, '提示', '请先选中一条记录', 'warning')
            return
        row = index.row()
        record_id = self.table_model.get_record_id(row)
        if record_id is None:
            show_message(self, '提示', '不能删除统计行', 'warning')
            return
        ok = confirm_action(self, '确认删除', '确定要删除这条记录吗？')
        if not ok:
            return
        if record_id > 0:
            self.pending_deletes.append(record_id)
            removed_edit = self.pending_edits.pop(record_id, None)
            all_db = get_all_records(None)
            deleted_row = None
            for r in all_db:
                if r[0] == record_id:
                    deleted_row = r
                    break
            self._push_undo({
                'type': 'delete',
                'record_id': record_id,
                'deleted_row': deleted_row,
                'removed_edit': removed_edit
            })
        else:
            idx = -record_id - 1
            if idx < len(self.pending_records):
                removed_data = self.pending_records.pop(idx)
                self._push_undo({
                    'type': 'delete_pending',
                    'index': idx,
                    'data': removed_data
                })
        self.redo_stack.clear()
        self._refresh_data()
        self._reset_auto_save_timer()

    def _reset_auto_save_timer(self):
        self.auto_save_timer.start(5 * 60 * 1000)

    def _auto_save(self):
        if not self.pending_records and not self.pending_deletes and not self.pending_edits:
            return
        self._perform_save(silent=True)
        self.status_label.setText('已自动保存')

    def _on_save(self):
        if not self.pending_records and not self.pending_deletes and not self.pending_edits:
            show_message(self, '提示', '没有需要保存的更改')
            return
        msg_parts = []
        if self.pending_records:
            msg_parts.append(f'新增 {len(self.pending_records)} 条')
        if self.pending_deletes:
            msg_parts.append(f'删除 {len(self.pending_deletes)} 条')
        if self.pending_edits:
            msg_parts.append(f'编辑 {len(self.pending_edits)} 条')
        ok = confirm_action(self, '确认保存', '\n'.join(msg_parts) + '\n\n确定保存吗？')
        if not ok:
            return
        self._perform_save(silent=False)
        self._reset_auto_save_timer()

    def _perform_save(self, silent=False):
        errors = []
        for rid in self.pending_deletes:
            try:
                delete_record(rid)
            except Exception as e:
                errors.append(f'删除失败: {e}')

        for rid, new_rec in self.pending_edits.items():
            ns_expr = new_rec[11] if len(new_rec) > 11 else ''
            nc_expr = new_rec[12] if len(new_rec) > 12 else ''
            hs_expr = new_rec[13] if len(new_rec) > 13 else ''
            hc_expr = new_rec[14] if len(new_rec) > 14 else ''
            try:
                update_record(rid, new_rec[1], '', '', new_rec[2], new_rec[3],
                              new_rec[4], new_rec[5], new_rec[6], new_rec[7],
                              new_rec[9] if len(new_rec) > 9 else '',
                              new_rec[10] if len(new_rec) > 10 else '',
                              ns_expr, nc_expr, hs_expr, hc_expr)
            except Exception as e:
                errors.append(f'编辑失败 (id={rid}): {e}')

        pending_exprs = {}
        for rec in self.table_model.records:
            rid = rec[0]
            if rid < 0:
                idx = -rid - 1
                if idx < len(self.pending_records):
                    pending_exprs[idx] = (
                        rec[11] if len(rec) > 11 else '',
                        rec[12] if len(rec) > 12 else '',
                        rec[13] if len(rec) > 13 else '',
                        rec[14] if len(rec) > 14 else ''
                    )
        new_id_map = {}
        for i, pr in enumerate(self.pending_records):
            ns_expr, nc_expr, hs_expr, hc_expr = pending_exprs.get(i, ('', '', '', ''))
            try:
                new_id = add_record(*pr,
                                    normal_salary_expr=ns_expr,
                                    normal_consume_expr=nc_expr,
                                    hero_salary_expr=hs_expr,
                                    hero_consume_expr=hc_expr)
                new_id_map[i] = new_id
            except Exception as e:
                errors.append(f'新增失败: {e}')

        new_undo = []
        for item in self.undo_stack:
            if item['type'] == 'add':
                idx = item['index']
                if idx in new_id_map:
                    new_id = new_id_map[idx]
                    with get_connection() as conn:
                        cur = conn.cursor()
                        cur.execute('SELECT * FROM salary_records WHERE id=?', (new_id,))
                        row = cur.fetchone()
                    if row:
                        new_undo.append({
                            'type': 'add_db',
                            'record_id': new_id,
                            'inserted_row': row
                        })
            elif item['type'] == 'edit':
                rid = item['record_id']
                if rid in self.pending_edits:
                    old = item['old_record']
                    new = item['new_record']
                    new_undo.append({
                        'type': 'edit_db',
                        'record_id': rid,
                        'old_record': old,
                        'new_record': new
                    })
            elif item['type'] == 'edit_pending':
                idx = item['pending_idx']
                if idx in new_id_map and idx < len(self.pending_records):
                    new_id = new_id_map[idx]
                    old_data = item['old_data']
                    new_data = item['new_data']
                    old_rec = (new_id, old_data[0], old_data[3], old_data[4], old_data[5], old_data[6], old_data[7], old_data[8], 0, old_data[9], old_data[10])
                    new_rec = (new_id, new_data[0], new_data[3], new_data[4], new_data[5], new_data[6], new_data[7], new_data[8], 0, new_data[9], new_data[10])
                    new_undo.append({
                        'type': 'edit_db',
                        'record_id': new_id,
                        'old_record': old_rec,
                        'new_record': new_rec
                    })
            elif item['type'] == 'delete':
                rid = item['record_id']
                new_undo.append({
                    'type': 'delete_db',
                    'record_id': rid,
                    'deleted_row': item['deleted_row']
                })
            elif item['type'] == 'delete_pending':
                pass

        self.undo_stack = new_undo
        self.redo_stack.clear()
        self.pending_records.clear()
        self.pending_deletes.clear()
        self.pending_edits.clear()

        if not silent:
            if errors:
                show_message(self, '部分失败', '\n'.join(errors), 'warning')
            else:
                show_message(self, '保存成功', '所有更改已保存')
        self._refresh_data()

    def _on_import(self):
        dialog = QFileDialog(self)
        dialog.setWindowTitle('选择Excel文件')
        dialog.setNameFilter('Excel文件 (*.xlsx *.xls)')
        dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        dialog.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        dialog.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        dialog.setStyleSheet(self.styleSheet())
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        if not os.path.exists(desktop):
            desktop = os.path.expanduser('~')
        dialog.setDirectory(desktop)
        if dialog.exec() != QFileDialog.DialogCode.Accepted:
            return
        file_paths = dialog.selectedFiles()
        if not file_paths:
            return
        file_path = file_paths[0]
        records, error = import_excel(file_path)
        if error:
            show_message(self, '导入失败', error, 'error')
            return
        if not records:
            show_message(self, '提示', '未发现有效数据')
            return
        if self.pending_records or self.pending_deletes or self.pending_edits:
            ok = confirm_action(self, '未保存的更改', '导入前将放弃当前未保存的更改，确定继续？')
            if not ok:
                return
            self.pending_records.clear()
            self.pending_deletes.clear()
            self.pending_edits.clear()
            self.undo_stack.clear()
            self.redo_stack.clear()
        count = 0
        for rec in records:
            try:
                add_record(*rec)
                count += 1
            except Exception as e:
                show_message(self, '错误', f'保存失败: {e}', 'warning')
        show_message(self, '导入成功', f'成功导入 {count} 条记录')
        self._refresh_data()

    def _on_export(self):
        data_records = [r for r in self.table_model.records if r[2] not in ('平均', '合计')]
        current_date = self.date_filter_combo.currentText()
        if not data_records:
            show_message(self, '提示', '当前没有数据可导出', 'warning')
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, '保存Excel文件', f'{current_date}.xlsx',
            'Excel文件 (*.xlsx)',
            options=QFileDialog.Option.DontUseNativeDialog
        )
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            if current_date and current_date != '全部':
                ws.cell(row=1, column=1, value=f'日期{current_date}')
            else:
                ws.cell(row=1, column=1, value='日期全部')
            headers = self.table_model.visible_columns
            for i, h in enumerate(headers, 1):
                ws.cell(row=3, column=i, value=h)
            for row_idx, record in enumerate(data_records, 4):
                for col_idx, col_name in enumerate(headers, 1):
                    idx = SalaryTableModel.FIELD_MAP[col_name]
                    value = record[idx]
                    if col_name in ('普通工资', '普通消费', '英雄工资', '英雄消费', '总工资'):
                        value = number_to_brick(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            wb.save(file_path)
            show_message(self, '导出成功', f'已保存到 {os.path.basename(file_path)}')
        except Exception as e:
            show_message(self, '导出失败', str(e), 'error')

    # ---------- 撤销 / 重做 ----------
    def _on_undo(self):
        if not self.undo_stack:
            self.status_label.setText('没有可撤销的操作')
            return
        item = self.undo_stack.pop()
        try:
            if item['type'] == 'edit_db':
                old = item['old_record']
                new_r = item['new_record']
                ns_expr = old[11] if len(old) > 11 else ''
                nc_expr = old[12] if len(old) > 12 else ''
                hs_expr = old[13] if len(old) > 13 else ''
                hc_expr = old[14] if len(old) > 14 else ''
                update_record(old[0], old[1], '', '', old[2], old[3],
                              old[4], old[5], old[6], old[7],
                              old[9] if len(old) > 9 else '',
                              old[10] if len(old) > 10 else '',
                              ns_expr, nc_expr, hs_expr, hc_expr)
                self.redo_stack.append({
                    'type': 'edit_db',
                    'record_id': old[0],
                    'old_record': new_r,
                    'new_record': old
                })
            elif item['type'] == 'add_db':
                delete_record(item['record_id'])
                self.redo_stack.append({
                    'type': 'add_db',
                    'record_id': item['record_id'],
                    'inserted_row': item['inserted_row']
                })
            elif item['type'] == 'delete_db':
                row = item['deleted_row']
                ns_expr = row[11] if len(row) > 11 else ''
                nc_expr = row[12] if len(row) > 12 else ''
                hs_expr = row[13] if len(row) > 13 else ''
                hc_expr = row[14] if len(row) > 14 else ''
                with get_connection() as conn:
                    conn.execute('''
                        INSERT INTO salary_records 
                        (id, date_range, start_month, start_day, end_month, end_day, sort_key,
                         character_name, faction,
                         normal_salary, normal_consume, hero_salary, hero_consume, total_salary,
                         team_mark, drop_info,
                         normal_salary_expr, normal_consume_expr, hero_salary_expr, hero_consume_expr)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ''', (row[0], row[1], row[2], row[3], row[4], row[5], row[6],
                          row[7], row[8], row[9], row[10], row[11], row[12], row[13],
                          row[14] if len(row) > 14 else '',
                          row[15] if len(row) > 15 else '',
                          ns_expr, nc_expr, hs_expr, hc_expr))
                    conn.commit()
                self.redo_stack.append({
                    'type': 'delete_db',
                    'record_id': row[0],
                    'deleted_row': row
                })
            elif item['type'] == 'edit':
                rid = item['record_id']
                old_rec = item['old_record']
                new_rec = item['new_record']
                self.pending_edits[rid] = old_rec
                self.redo_stack.append({'type': 'edit', 'record_id': rid, 'old_record': new_rec, 'new_record': old_rec})
            elif item['type'] == 'edit_pending':
                idx = item['pending_idx']
                if idx < len(self.pending_records):
                    self.pending_records[idx] = item['old_data']
                self.redo_stack.append({'type': 'edit_pending', 'pending_idx': idx, 'old_data': item['new_data'], 'new_data': item['old_data']})
            elif item['type'] == 'add':
                idx = item['index']
                if idx < len(self.pending_records):
                    removed = self.pending_records.pop(idx)
                    self.redo_stack.append({'type': 'add', 'index': idx, 'data': removed})
            elif item['type'] == 'delete':
                rid = item['record_id']
                if rid in self.pending_deletes:
                    self.pending_deletes.remove(rid)
                if item.get('removed_edit'):
                    self.pending_edits[rid] = item['removed_edit']
                self.redo_stack.append({'type': 'delete', 'record_id': rid, 'deleted_row': item['deleted_row'], 'removed_edit': item.get('removed_edit')})
            elif item['type'] == 'delete_pending':
                idx = item['index']
                self.pending_records.insert(idx, item['data'])
                self.redo_stack.append({'type': 'delete_pending', 'index': idx, 'data': item['data']})
            self._refresh_data()
            self.status_label.setText('已撤销')
        except Exception as e:
            self.status_label.setText(f'撤销失败: {e}')
            self.undo_stack.append(item)

    def _on_redo(self):
        if not self.redo_stack:
            self.status_label.setText('没有可重做的操作')
            return
        item = self.redo_stack.pop()
        try:
            if item['type'] == 'edit_db':
                new_r = item['old_record']
                ns_expr = new_r[11] if len(new_r) > 11 else ''
                nc_expr = new_r[12] if len(new_r) > 12 else ''
                hs_expr = new_r[13] if len(new_r) > 13 else ''
                hc_expr = new_r[14] if len(new_r) > 14 else ''
                update_record(new_r[0], new_r[1], '', '', new_r[2], new_r[3],
                              new_r[4], new_r[5], new_r[6], new_r[7],
                              new_r[9] if len(new_r) > 9 else '',
                              new_r[10] if len(new_r) > 10 else '',
                              ns_expr, nc_expr, hs_expr, hc_expr)
                self._push_undo({'type': 'edit_db', 'record_id': new_r[0], 'old_record': item['new_record'], 'new_record': new_r})
            elif item['type'] == 'add_db':
                row = item['inserted_row']
                add_record(row[1], '', '', row[7], row[8],
                           row[9], row[10], row[11], row[12],
                           row[14] if len(row) > 14 else '',
                           row[15] if len(row) > 15 else '',
                           normal_salary_expr=row[11] if len(row)>11 else '',
                           normal_consume_expr=row[12] if len(row)>12 else '',
                           hero_salary_expr=row[13] if len(row)>13 else '',
                           hero_consume_expr=row[14] if len(row)>14 else '')
                self._push_undo({'type': 'add_db', 'record_id': row[0], 'inserted_row': row})
            elif item['type'] == 'delete_db':
                row = item['deleted_row']
                delete_record(row[0])
                self._push_undo({'type': 'delete_db', 'record_id': row[0], 'deleted_row': row})
            elif item['type'] == 'edit':
                rid = item['record_id']
                self.pending_edits[rid] = item['old_record']
                self._push_undo({'type': 'edit', 'record_id': rid, 'old_record': item['new_record'], 'new_record': item['old_record']})
            elif item['type'] == 'edit_pending':
                idx = item['pending_idx']
                if idx < len(self.pending_records):
                    self.pending_records[idx] = item['old_data']
                self._push_undo({'type': 'edit_pending', 'pending_idx': idx, 'old_data': item['new_data'], 'new_data': item['old_data']})
            elif item['type'] == 'add':
                idx = item['index']
                self.pending_records.insert(idx, item['data'])
                self._push_undo({'type': 'add', 'index': idx, 'data': item['data']})
            elif item['type'] == 'delete':
                rid = item['record_id']
                self.pending_deletes.append(rid)
                self._push_undo({'type': 'delete', 'record_id': rid, 'deleted_row': item['deleted_row'], 'removed_edit': item.get('removed_edit')})
            elif item['type'] == 'delete_pending':
                idx = item['index']
                if idx < len(self.pending_records):
                    self.pending_records.pop(idx)
                self._push_undo({'type': 'delete_pending', 'index': idx, 'data': item['data']})
            self._refresh_data()
            self.status_label.setText('已重做')
        except Exception as e:
            self.status_label.setText(f'重做失败: {e}')
            self.redo_stack.append(item)

    # ---------- 右键菜单 ----------
    def _on_table_context_menu(self, pos: QPoint):
        menu = QMenu(self)
        gray_action = QAction('运算式行变灰', menu, checkable=True)
        gray_action.setChecked(self.gray_expression_rows)
        gray_action.toggled.connect(self._toggle_gray_rows)
        menu.addAction(gray_action)

        menu.addSeparator()

        index = self.table_view.indexAt(pos)
        row = index.row() if index.isValid() else -1
        if row >= 0:
            delete_action = QAction('删除该行', menu)
            delete_action.triggered.connect(lambda: self._delete_row_at(row))
            menu.addAction(delete_action)

        menu.exec(self.table_view.viewport().mapToGlobal(pos))

    def _toggle_gray_rows(self, enabled):
        self.gray_expression_rows = enabled
        self.table_model.set_gray_expression_rows(enabled)
        self._save_settings()

    def _delete_row_at(self, row):
        self.table_view.selectRow(row)
        self._on_delete()
